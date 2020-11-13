#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Nov 10 13:01:53 2020

@author: wangmeiqi
"""

import os

import SimpleITK as sitk

#tools for HCC


#read dictionary and files
filenames = []

for root, dir, files in os.walk(route):
    
    for file in files:
       
        if file != '.DS_Store':
           
            filenames.append(file)
            
#end
            

#simpleitk read and ourput image
read_img = sitk.ReadImage(img_path)

img_array = sitk.GetArrayFromImage(read_img)

output_img = sitk.GetImageFromArray(output_img)

output_img.CopyInformation(read_img)

sitk.WriteImage(output_img, save_path)

#end


#print headers
data_dir = ' '

sitkReader = sitk.ImageFileReader()

sitkReader.SetFileName(data_dir)

sitkImage = sitkReader.Execute()


for key in sitkReader.GetMetaDataKeys():
    
  value = sitkReader.GetMetaData(key)
  
  print("Key:{} : {}".format(key, value))
  

#excel
from openpyxl import Workbook

wb = Workbook()

ws = wb.active

ws['A1'] = " "

ws['B1'] = " "

ws['C1'] = " "

ws['A'+str(row)] = " "

ws['B'+str(row)] = " "

ws['C'+str(row)] = " "

wb.save(save_path)

#end


#rename
main_route = " " #here

file = " " #here

path = os.path.join(main_route, file)


for fname in os.listdir(path):
    
    idx = fname.find(" ") #here
    
    new_fname = fname[idx:]
    
    print(os.path.join(path, fname))
    
    print(os.path.join(path, new_fname))
    
    os.rename(os.path.join(path, fname), os.path.join(path, new_fname))
    
#end


#auto mkdir
num = len(dir_)

for i in range(num):
    
    path = os.path.join(save_route, dir_[i])
    
    if not os.path.isdir(path):
        
        os.mkdir(path)
        
#end
        

#刪除指定檔案
route = ' '

for root, dirs, files in os.walk(route):

    for file in files:
        
        if file[:6] == 'post_v': #here
            
            filename = os.path.join(root, file)
            
            os.remove(filename)
            
            print(filename)
            
#end
            

#copy file
import shutil

src = "from where"

dst = "to where"

shutil.copyfile(src, dst)

